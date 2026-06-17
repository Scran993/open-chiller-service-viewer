#!/usr/bin/env python3
"""
Guarded Modbus RTU writer for Danfoss/KTK MCX service work.

Default behaviour is read-only. Actual write commands require --enable-write
and an interactive confirmation. Dangerous parameters also require the phrase
I UNDERSTAND.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_LOG = ROOT / "write-logs" / "write-attempts.csv"
DEFAULT_BACKUP_DIR = ROOT / "backups"

READ_ONLY_WORDS = {
    "read",
    "r",
    "ro",
    "readonly",
    "read only",
}

WRITE_WORDS = {
    "rw",
    "r/w",
    "write",
    "read/write",
    "read write",
}

HARD_BLOCK_TERMS = (
    "alarm",
    "fault",
    "status",
    "actual",
    "sensor",
    "probe",
    "pressure",
    "temperature",
    "rpm",
    "speed",
    "current",
    "voltage",
    "power",
    "state",
    "notification",
    "hours",
)

DANGEROUS_TERMS = (
    "compressor",
    "valve",
    "liquid",
    "pump",
    "fan",
    "enable",
    "control",
    "config",
    "setpoint",
    "capacity",
    "limit",
)


class SafeWriteError(Exception):
    pass


@dataclass
class Point:
    label: str
    description: str = ""
    section: str = ""
    register: int | None = None
    function: str = "03"
    access: str = ""
    minimum: float | None = None
    maximum: float | None = None
    value_type: str = ""
    unit: str = ""
    scale: float = 1.0
    enum: dict[int, str] = field(default_factory=dict)
    notes: str = ""

    @property
    def display_name(self) -> str:
        return self.label or self.description or f"register {self.register}"

    @property
    def is_rw(self) -> bool:
        access = normalize_access(self.access)
        return access in WRITE_WORDS

    @property
    def is_read_only(self) -> bool:
        access = normalize_access(self.access)
        return access in READ_ONLY_WORDS or access not in WRITE_WORDS

    @property
    def is_dangerous(self) -> bool:
        text = " ".join([self.label, self.description, self.section, self.notes]).lower()
        return any(term in text for term in DANGEROUS_TERMS)

    @property
    def is_hard_blocked(self) -> bool:
        text = " ".join([self.label, self.description, self.section, self.notes]).lower()
        return any(term in text for term in HARD_BLOCK_TERMS)


def normalize_access(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text or text.lower() in {"-", "---", "none", "nan"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_int(value: Any) -> int | None:
    number = parse_number(value)
    if number is None:
        return None
    if not math.isfinite(number):
        return None
    return int(number)


def parse_enum(text: str) -> dict[int, str]:
    out: dict[int, str] = {}
    if not text:
        return out
    for match in re.finditer(r"(-?\d+)\s*[-=:]\s*([A-Za-z][A-Za-z0-9_ /.-]*)", text):
        out[int(match.group(1))] = match.group(2).strip()
    return out


def infer_header(row: list[Any]) -> list[str]:
    return [str(cell or "").strip().lower() for cell in row]


def pick(row: dict[str, Any], *names: str) -> Any:
    lowered = {k.strip().lower(): v for k, v in row.items()}
    for name in names:
        if name in lowered and lowered[name] not in (None, ""):
            return lowered[name]
    return ""


def point_from_row(row: dict[str, Any]) -> Point | None:
    label = str(pick(row, "label", "name", "variable", "tag", "id", "parameter")).strip()
    description = str(pick(row, "description", "desc", "notes", "comment")).strip()
    register = parse_int(pick(row, "adu", "register", "address", "modbus", "adu/register"))
    if not label and not description:
        return None
    if register is None:
        return None
    scale = parse_number(pick(row, "scale", "gain", "multiplier"))
    enum_source = pick(row, "enum", "options")
    value_type = str(pick(row, "value_type", "value/type", "value", "type")).strip()
    enum: dict[int, str]
    if isinstance(enum_source, dict):
        enum = {int(k): str(v) for k, v in enum_source.items()}
        enum_text = value_type
    else:
        enum_text = str(enum_source or value_type).strip()
        enum = parse_enum(enum_text)
    return Point(
        label=label or description,
        description=description,
        section=str(pick(row, "section", "group", "menu")).strip(),
        register=register,
        function=str(pick(row, "function", "function code", "fc") or "03").replace("0x", "").zfill(2),
        access=str(pick(row, "access", "rw", "r/w", "read_write", "read/write")).strip(),
        minimum=parse_number(pick(row, "min", "minimum", "low")),
        maximum=parse_number(pick(row, "max", "maximum", "high")),
        value_type=value_type or enum_text,
        unit=str(pick(row, "unit", "units", "range")).strip(),
        scale=scale if scale not in (None, 0) else 1.0,
        enum=enum,
        notes=str(pick(row, "notes", "comment")).strip(),
    )


def load_json_points(path: Path) -> tuple[str, list[Point]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    name = data.get("name") or data.get("profile") or path.stem
    rows = data.get("registers") or data.get("points") or []
    points: list[Point] = []
    for row in rows:
        point = point_from_row(row)
        if point:
            points.append(point)
    return name, points


def load_csv_points(path: Path) -> tuple[str, list[Point]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        points = [p for p in (point_from_row(row) for row in csv.DictReader(handle)) if p]
    return path.stem, points


def load_xlsx_points(path: Path) -> tuple[str, list[Point]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SafeWriteError("openpyxl is required to import XLSX files") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_index = 0
    for idx, row in enumerate(rows[:30]):
        headers = infer_header(list(row))
        if any(h in headers for h in ("label", "name", "description", "adu", "register", "address")):
            header_index = idx
            break
    headers = infer_header(list(rows[header_index]))
    points: list[Point] = []
    for row in rows[header_index + 1 :]:
        data = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
        point = point_from_row(data)
        if point:
            points.append(point)
    return path.stem, points


def load_text_points(path: Path) -> tuple[str, list[Point]]:
    lines = [line.strip() for line in path.read_text(encoding="utf-8", errors="ignore").splitlines()]
    lines = [line for line in lines if line]
    section = ""
    block_start = 0
    points: list[Point] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if ">" in line and line.lower() not in WRITE_WORDS and normalize_access(line) not in READ_ONLY_WORDS:
            section = line
            block_start = i + 1
            i += 1
            continue
        access = normalize_access(line)
        if access in WRITE_WORDS | READ_ONLY_WORDS and i + 1 < len(lines) and parse_int(lines[i + 1]) is not None:
            block = lines[block_start:i]
            if len(block) >= 4:
                unit = ""
                value_type = block[4] if len(block) > 4 else ""
                if len(block) > 5 and not block[5].lower().startswith("enum"):
                    unit = block[5]
                enum_text = value_type if re.search(r"\d+\s*[-=:]", value_type) else ""
                points.append(
                    Point(
                        label=block[0],
                        description=block[1] if len(block) > 1 else "",
                        section=section,
                        minimum=parse_number(block[2] if len(block) > 2 else ""),
                        maximum=parse_number(block[3] if len(block) > 3 else ""),
                        value_type=value_type,
                        unit=unit,
                        access=line,
                        register=parse_int(lines[i + 1]),
                        enum=parse_enum(enum_text),
                    )
                )
            block_start = i + 2
            i += 2
            continue
        i += 1
    return path.stem, points


def load_points(path: Path) -> tuple[str, list[Point]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return load_json_points(path)
    if suffix == ".csv":
        return load_csv_points(path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_xlsx_points(path)
    if suffix in {".txt", ".log"}:
        return load_text_points(path)
    raise SafeWriteError(f"Unsupported profile/list type: {path.suffix}")


def find_point(points: list[Point], label: str) -> Point:
    wanted = label.strip().lower()
    exact = [p for p in points if p.label.lower() == wanted]
    if exact:
        return exact[0]
    loose = [p for p in points if wanted in p.label.lower() or wanted in p.description.lower()]
    if len(loose) == 1:
        return loose[0]
    if loose:
        names = ", ".join(p.label for p in loose[:12])
        raise SafeWriteError(f"More than one point matches {label!r}: {names}")
    raise SafeWriteError(f"No point found for {label!r}")


def require_serial():
    try:
        import serial  # type: ignore
    except ImportError as exc:
        raise SafeWriteError("pyserial is not installed. Dry-run/import works, but COM-port read/write needs pyserial.") from exc
    return serial


def crc16(data: bytes) -> int:
    crc = 0xFFFF
    for byte in data:
        crc ^= byte
        for _ in range(8):
            if crc & 1:
                crc = (crc >> 1) ^ 0xA001
            else:
                crc >>= 1
    return crc & 0xFFFF


def add_crc(frame: bytes) -> bytes:
    crc = crc16(frame)
    return frame + bytes([crc & 0xFF, crc >> 8])


def check_crc(frame: bytes) -> None:
    if len(frame) < 4:
        raise SafeWriteError("Short Modbus response")
    got = frame[-2] | (frame[-1] << 8)
    want = crc16(frame[:-2])
    if got != want:
        raise SafeWriteError(f"CRC mismatch: got {got:04X}, expected {want:04X}")


class ModbusRtu:
    def __init__(self, port: str, baud: int, timeout: float):
        serial = require_serial()
        self.serial = serial.Serial(
            port=port,
            baudrate=baud,
            bytesize=8,
            parity=serial.PARITY_NONE,
            stopbits=1,
            timeout=timeout,
            write_timeout=timeout,
        )

    def close(self) -> None:
        self.serial.close()

    def transact(self, request: bytes, expected: int, timeout: float) -> bytes:
        self.serial.reset_input_buffer()
        self.serial.write(add_crc(request))
        self.serial.flush()
        deadline = time.time() + timeout
        response = b""
        while len(response) < expected and time.time() < deadline:
            chunk = self.serial.read(expected - len(response))
            if chunk:
                response += chunk
        check_crc(response)
        if response[1] & 0x80:
            raise SafeWriteError(f"Modbus exception code {response[2]}")
        return response

    def read_holding(self, slave: int, address: int, quantity: int, timeout: float) -> list[int]:
        request = bytes([slave, 0x03, address >> 8, address & 0xFF, quantity >> 8, quantity & 0xFF])
        expected = 5 + quantity * 2
        response = self.transact(request, expected, timeout)
        if response[1] != 0x03:
            raise SafeWriteError(f"Unexpected function in response: {response[1]}")
        return [(response[3 + i * 2] << 8) | response[4 + i * 2] for i in range(quantity)]

    def write_single(self, slave: int, address: int, value: int, timeout: float) -> None:
        request = bytes([slave, 0x06, address >> 8, address & 0xFF, value >> 8, value & 0xFF])
        response = self.transact(request, 8, timeout)
        if response[:-2] != request:
            raise SafeWriteError("Write response did not echo the request")

    def write_multiple(self, slave: int, address: int, values: list[int], timeout: float) -> None:
        qty = len(values)
        body = bytearray([slave, 0x10, address >> 8, address & 0xFF, qty >> 8, qty & 0xFF, qty * 2])
        for value in values:
            body.extend([value >> 8, value & 0xFF])
        response = self.transact(bytes(body), 8, timeout)
        if response[1] != 0x10:
            raise SafeWriteError(f"Unexpected function in response: {response[1]}")


def wire_address(point: Point, offset: int) -> int:
    if point.register is None:
        raise SafeWriteError("Point has no ADU/register")
    address = point.register + offset
    if not 0 <= address <= 0xFFFF:
        raise SafeWriteError(f"Register {point.register} with offset {offset} is outside Modbus range")
    return address


def raw_to_display(point: Point, raw: int | None) -> str:
    if raw is None:
        return "--"
    if raw in point.enum:
        return f"{raw} - {point.enum[raw]}"
    value = raw * point.scale
    if point.scale != 1:
        return f"{value:g}"
    return str(raw)


def parse_proposed_value(point: Point, value_text: str, raw_value: bool) -> tuple[int, str]:
    text = value_text.strip()
    if not text:
        raise SafeWriteError("No proposed value supplied")
    if point.enum:
        for number, name in point.enum.items():
            if text.lower() == name.lower() or text.lower() == f"{number} - {name}".lower():
                return number, f"{number} - {name}"
    if text.lower() in {"on", "true", "yes"}:
        text = "1"
    elif text.lower() in {"off", "false", "no"}:
        text = "0"
    number = parse_number(text)
    if number is None:
        raise SafeWriteError(f"{value_text!r} is not a valid numeric or enum value")
    raw = int(round(number if raw_value or point.scale == 1 else number / point.scale))
    if not 0 <= raw <= 0xFFFF:
        raise SafeWriteError(f"Raw value {raw} is outside 0..65535")
    display = raw_to_display(point, raw)
    return raw, display


def validate_point_for_write(point: Point, raw: int, advanced_manual: bool) -> list[str]:
    warnings: list[str] = []
    if point.is_read_only:
        raise SafeWriteError(f"{point.display_name} is not marked RW. Access is {point.access!r}; write blocked.")
    if point.is_hard_blocked and not advanced_manual:
        raise SafeWriteError(f"{point.display_name} looks like live/alarm/status data. Write blocked.")
    engineering = raw * point.scale
    if point.minimum is not None and engineering < point.minimum:
        raise SafeWriteError(f"Proposed value {engineering:g} is below minimum {point.minimum:g}")
    if point.maximum is not None and engineering > point.maximum:
        raise SafeWriteError(f"Proposed value {engineering:g} is above maximum {point.maximum:g}")
    if point.enum and raw not in point.enum:
        allowed = ", ".join(f"{k}={v}" for k, v in sorted(point.enum.items()))
        raise SafeWriteError(f"Proposed enum {raw} is not one of: {allowed}")
    if point.minimum is not None and point.maximum is not None and not (point.minimum <= engineering <= point.maximum):
        warnings.append("Proposed value is outside the published range.")
    return warnings


def print_write_preview(point: Point, current_raw: int | None, proposed_raw: int, proposed_display: str, offset: int, dry_run: bool) -> None:
    print("\nCONTROLLED WRITE PREVIEW")
    print("Writing live HVAC controller parameters can cause compressor trips, valve movement, pump/fan changes, or controller reboot.")
    print(f"Label:          {point.label}")
    print(f"Description:    {point.description}")
    print(f"Section:        {point.section}")
    print(f"ADU/register:   {point.register}  (wire address {wire_address(point, offset)} with offset {offset})")
    print(f"Access:         {point.access}")
    print(f"Current value:  {raw_to_display(point, current_raw)}  raw={current_raw if current_raw is not None else '--'}")
    print(f"New value:      {proposed_display}  raw={proposed_raw}")
    print(f"Unit:           {point.unit or '-'}")
    print(f"Min/max:        {point.minimum if point.minimum is not None else '-'} / {point.maximum if point.maximum is not None else '-'}")
    if point.enum:
        print("Enum options:   " + ", ".join(f"{k}={v}" for k, v in sorted(point.enum.items())))
    if point.label.lower() == "lv1":
        print("lv1 warning: changing this may alter liquid valve control strategy and refrigerant feed.")
    if dry_run:
        print("Dry-run:        no Modbus write will be sent.")


def prompt_confirmation(point: Point, proposed_display: str) -> None:
    answer = input(f"\nType WRITE {point.label} to continue: ").strip()
    if answer != f"WRITE {point.label}":
        raise SafeWriteError("Confirmation did not match; write cancelled.")
    if point.is_dangerous:
        phrase = input('Dangerous parameter. Type I UNDERSTAND to continue: ').strip()
        if phrase != "I UNDERSTAND":
            raise SafeWriteError("Safety phrase did not match; write cancelled.")
    final = input(f"Final check. Type the proposed value exactly ({proposed_display}): ").strip()
    if final != proposed_display:
        raise SafeWriteError("Final value confirmation did not match; write cancelled.")


def append_write_log(path: Path, row: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    exists = path.exists()
    fields = [
        "timestamp",
        "device_address",
        "register",
        "wire_address",
        "label",
        "old_raw",
        "old_value",
        "new_raw",
        "new_value",
        "success",
        "dry_run",
        "exception",
    ]
    with path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        if not exists:
            writer.writeheader()
        writer.writerow({field: row.get(field, "") for field in fields})


def open_bus(args: argparse.Namespace) -> ModbusRtu:
    return ModbusRtu(args.port, args.baud, args.timeout)


def cmd_import(args: argparse.Namespace) -> int:
    name, points = load_points(Path(args.source))
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    data = {
        "name": args.name or name,
        "warning": "This profile may contain RW parameters. Use safe_modbus_writer.py; the viewer remains read-only.",
        "registers": [
            {
                "section": p.section,
                "label": p.label,
                "name": p.label,
                "description": p.description,
                "register": p.register,
                "function": p.function,
                "access": p.access,
                "min": p.minimum,
                "max": p.maximum,
                "value_type": p.value_type,
                "unit": p.unit,
                "scale": p.scale,
                "enum": p.enum,
                "notes": p.notes,
            }
            for p in points
        ],
    }
    out.write_text(json.dumps(data, indent=2), encoding="utf-8")
    rw = sum(1 for p in points if p.is_rw)
    print(f"Imported {len(points)} points ({rw} RW) to {out}")
    return 0


def cmd_inspect(args: argparse.Namespace) -> int:
    _, points = load_points(Path(args.profile))
    point = find_point(points, args.label)
    print(json.dumps(point.__dict__, indent=2, default=str))
    if point.is_read_only:
        print("Write status: BLOCKED unless the source list marks this point RW.")
    elif point.is_hard_blocked:
        print("Write status: BLOCKED as live/alarm/status/sensor-style data.")
    else:
        print("Write status: RW candidate, still requires --enable-write and confirmation.")
    return 0


def cmd_read(args: argparse.Namespace) -> int:
    _, points = load_points(Path(args.profile))
    point = find_point(points, args.label)
    bus = open_bus(args)
    try:
        raw = bus.read_holding(args.slave, wire_address(point, args.offset), 1, args.timeout)[0]
    finally:
        bus.close()
    print(f"{point.label}: {raw_to_display(point, raw)} raw={raw} register={point.register} offset={args.offset}")
    return 0


def cmd_backup(args: argparse.Namespace) -> int:
    name, points = load_points(Path(args.profile))
    rw_points = [p for p in points if p.is_rw]
    if not rw_points:
        raise SafeWriteError("No safe RW points found to back up.")
    DEFAULT_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    out = Path(args.out) if args.out else DEFAULT_BACKUP_DIR / f"rw-backup-{dt.datetime.now():%Y%m%d-%H%M%S}.json"
    bus = open_bus(args)
    rows = []
    try:
        for point in rw_points:
            try:
                raw = bus.read_holding(args.slave, wire_address(point, args.offset), 1, args.timeout)[0]
                error = ""
            except Exception as exc:  # noqa: BLE001 - logged for field diagnostics
                raw = None
                error = str(exc)
            rows.append(
                {
                    "label": point.label,
                    "description": point.description,
                    "register": point.register,
                    "raw": raw,
                    "value": raw_to_display(point, raw),
                    "error": error,
                }
            )
    finally:
        bus.close()
    data = {
        "timestamp": dt.datetime.now().isoformat(timespec="seconds"),
        "profile": name,
        "slave": args.slave,
        "baud": args.baud,
        "offset": args.offset,
        "rw_values": rows,
    }
    out.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(f"Saved RW backup to {out} ({len(rows)} points)")
    return 0


def cmd_write(args: argparse.Namespace) -> int:
    _, points = load_points(Path(args.profile))
    point = find_point(points, args.label)
    proposed_raw, proposed_display = parse_proposed_value(point, args.value, args.raw_value)
    validate_point_for_write(point, proposed_raw, args.advanced_manual)

    current_raw = None
    error = ""
    if not args.dry_run:
        bus = open_bus(args)
        try:
            current_raw = bus.read_holding(args.slave, wire_address(point, args.offset), 1, args.timeout)[0]
        finally:
            bus.close()

    print_write_preview(point, current_raw, proposed_raw, proposed_display, args.offset, args.dry_run)
    if args.good_value:
        good_raw, good_display = parse_proposed_value(point, args.good_value, args.raw_value)
        print(f"Known-good value: {good_display} raw={good_raw}")

    success = False
    try:
        if not args.dry_run:
            if not args.enable_write:
                raise SafeWriteError("Write mode is disabled. Re-run with --enable-write after reviewing the preview.")
            prompt_confirmation(point, proposed_display)
            bus = open_bus(args)
            try:
                if args.write_function == "16":
                    bus.write_multiple(args.slave, wire_address(point, args.offset), [proposed_raw], args.timeout)
                else:
                    bus.write_single(args.slave, wire_address(point, args.offset), proposed_raw, args.timeout)
                verify_raw = bus.read_holding(args.slave, wire_address(point, args.offset), 1, args.timeout)[0]
            finally:
                bus.close()
            success = verify_raw == proposed_raw
            print(f"Post-write readback: {raw_to_display(point, verify_raw)} raw={verify_raw}")
            if not success:
                raise SafeWriteError("Readback did not match the proposed value.")
        else:
            success = True
    except Exception as exc:  # noqa: BLE001 - this is the audit trail
        error = str(exc)
        print(f"Write result: FAILED - {error}")
        if not args.dry_run:
            append_write_log(args.log, log_row(args, point, current_raw, proposed_raw, proposed_display, False, error))
        return 2

    append_write_log(args.log, log_row(args, point, current_raw, proposed_raw, proposed_display, success, error))
    print("Write result: dry-run passed." if args.dry_run else "Write result: success.")
    return 0


def log_row(args: argparse.Namespace, point: Point, current_raw: int | None, proposed_raw: int, proposed_display: str, success: bool, error: str) -> dict[str, Any]:
    return {
        "timestamp": dt.datetime.now().isoformat(timespec="seconds"),
        "device_address": args.slave,
        "register": point.register,
        "wire_address": wire_address(point, args.offset),
        "label": point.label,
        "old_raw": current_raw if current_raw is not None else "",
        "old_value": raw_to_display(point, current_raw),
        "new_raw": proposed_raw,
        "new_value": proposed_display,
        "success": success,
        "dry_run": args.dry_run,
        "exception": error,
    }


def cmd_restore(args: argparse.Namespace) -> int:
    if not args.enable_write and not args.dry_run:
        raise SafeWriteError("Restore requires --enable-write, or use --dry-run.")
    _, points = load_points(Path(args.profile))
    by_key = {(p.label.lower(), p.register): p for p in points}
    backup = json.loads(Path(args.backup).read_text(encoding="utf-8"))
    rows = backup.get("rw_values") or []
    bus = None if args.dry_run else open_bus(args)
    try:
        for row in rows:
            key = (str(row.get("label", "")).lower(), parse_int(row.get("register")))
            point = by_key.get(key)
            raw = parse_int(row.get("raw"))
            if not point or raw is None:
                continue
            validate_point_for_write(point, raw, args.advanced_manual)
            print_write_preview(point, None, raw, raw_to_display(point, raw), args.offset, args.dry_run)
            if not args.batch_confirm:
                prompt_confirmation(point, raw_to_display(point, raw))
            elif args.batch_confirm != "I UNDERSTAND":
                raise SafeWriteError("Batch restore confirmation must be exactly I UNDERSTAND")
            if bus:
                bus.write_single(args.slave, wire_address(point, args.offset), raw, args.timeout)
            append_write_log(args.log, log_row(args, point, None, raw, raw_to_display(point, raw), True, ""))
    finally:
        if bus:
            bus.close()
    print("Restore dry-run complete." if args.dry_run else "Restore complete.")
    return 0


def add_connection_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--port", default="COM7")
    parser.add_argument("--baud", type=int, default=19200)
    parser.add_argument("--slave", type=int, required=True)
    parser.add_argument("--offset", type=int, default=-1, help="Use -1 for one-based ADU lists, 0 for zero-based lists.")
    parser.add_argument("--timeout", type=float, default=3.0)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Safe Modbus RW parameter tool for MCX controllers.")
    sub = parser.add_subparsers(dest="command", required=True)

    p_import = sub.add_parser("import", help="Convert CSV/XLSX/TXT/JSON variable list to guarded JSON.")
    p_import.add_argument("--source", required=True)
    p_import.add_argument("--out", required=True)
    p_import.add_argument("--name")
    p_import.set_defaults(func=cmd_import)

    p_inspect = sub.add_parser("inspect", help="Show metadata and write eligibility for one point.")
    p_inspect.add_argument("--profile", required=True)
    p_inspect.add_argument("--label", required=True)
    p_inspect.set_defaults(func=cmd_inspect)

    p_read = sub.add_parser("read", help="Read one holding register by label.")
    p_read.add_argument("--profile", required=True)
    p_read.add_argument("--label", required=True)
    add_connection_args(p_read)
    p_read.set_defaults(func=cmd_read)

    p_backup = sub.add_parser("backup", help="Read and save all safe RW parameter values.")
    p_backup.add_argument("--profile", required=True)
    p_backup.add_argument("--out")
    add_connection_args(p_backup)
    p_backup.set_defaults(func=cmd_backup)

    p_write = sub.add_parser("write", help="Validate, preview, and optionally write one RW register with FC06.")
    p_write.add_argument("--profile", required=True)
    p_write.add_argument("--label", required=True)
    p_write.add_argument("--value", required=True)
    p_write.add_argument("--good-value", help="Optional known-good value to display for comparison.")
    p_write.add_argument("--raw-value", action="store_true", help="Treat --value as raw register units.")
    p_write.add_argument("--dry-run", action="store_true", help="Validate and preview only; do not send a write.")
    p_write.add_argument("--enable-write", action="store_true", help="Required for any real Modbus write.")
    p_write.add_argument("--advanced-manual", action="store_true", help="Permit expert-only live/status-looking RW points.")
    p_write.add_argument("--write-function", choices=("06", "16"), default="06", help="FC06 is default. FC16 is optional and must be selected explicitly.")
    p_write.add_argument("--log", type=Path, default=DEFAULT_LOG)
    add_connection_args(p_write)
    p_write.set_defaults(func=cmd_write)

    p_restore = sub.add_parser("restore", help="Restore RW values from a backup file.")
    p_restore.add_argument("--profile", required=True)
    p_restore.add_argument("--backup", required=True)
    p_restore.add_argument("--dry-run", action="store_true")
    p_restore.add_argument("--enable-write", action="store_true")
    p_restore.add_argument("--batch-confirm", help="Use I UNDERSTAND to restore without per-point prompts.")
    p_restore.add_argument("--advanced-manual", action="store_true")
    p_restore.add_argument("--log", type=Path, default=DEFAULT_LOG)
    add_connection_args(p_restore)
    p_restore.set_defaults(func=cmd_restore)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return args.func(args)
    except SafeWriteError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
